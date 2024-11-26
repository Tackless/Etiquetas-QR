import qrcode
import openpyxl
from openpyxl.drawing.image import Image
import os

# Ruta del archivo de Excel y la carpeta para imágenes QR
excel_file = "códigos_qr.xlsx"
ruta_carpeta_imagenes = "C:\\Users\\Falcon\\Documents\\Falcon\\Scripts\\Etiquetas\\img"

# Verifica y crea todas las carpetas de la ruta si no existen
os.makedirs(ruta_carpeta_imagenes, exist_ok=True)

# Elige la celda en la que va iniciar (Del Drive)
celda_num_inicial = int(input("Ingresa en que celda vas a empezar: "))
print(f"Elegiste la celda {celda_num_inicial}.")

# Elige cuantas celdas se van a crear
celda_num_tamanio = int(input("Ingresa cuantas celdas vas a crear: "))
print(f"Vas a crear {celda_num_tamanio} celdas.")

# Cargar o crear el archivo de Excel
try:
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=1).value = "Dato"
    sheet.cell(row=1, column=2).value = "Código QR"

# Lista de datos para generar los códigos QR
datos = []

for celda in range(celda_num_inicial, celda_num_inicial + celda_num_tamanio):
    # comment: 
    link = "https://docs.google.com/spreadsheets/d/1nAdbMEj2fq2frxT42GbnS0p0JjGRxp-6/edit#gid=1326930466&range=E"
    datos.append(link+str(celda))
# end for

# Lista para almacenar las rutas de las imágenes que se generen
imagenes_generadas = []

# Generar y agregar códigos QR en Excel
for idx, dato in enumerate(datos, start=2):
    # Genera el código QR y guarda la imagen en la carpeta
    qr = qrcode.make(dato)
    temp_img = os.path.join(ruta_carpeta_imagenes, f"qr_{celda_num_inicial + (idx - 2)}.png")
    # f_img = os.path.join(ruta_carpeta_imagenes, f"f.png")
    qr.save(temp_img)
    imagenes_generadas.append(temp_img)  # Agrega la imagen a la lista

    sheet.row_dimensions[idx].height = 100
    sheet.column_dimensions['A'].width = 24.09
    sheet.column_dimensions['B'].width = 24.09
    
    # Inserta el dato y la imagen QR en Excel
    sheet.cell(row=idx, column=1).value = (celda_num_inicial + (idx - 2))
    img = Image(temp_img)
    img.width, img.height = 100, 100  # Ajusta tamaño si es necesario
    sheet.add_image(img, f"B{idx}")
    # img = Image(f_img)
    # img.width, img.height = 100, 100  # Ajusta tamaño si es necesario
    # sheet.add_image(img, f"B{idx}")



# Guardar cambios en el archivo de Excel
workbook.save(excel_file)
print(f"El archivo {excel_file} ha sido creado o actualizado con los códigos QR.")

# Ahora, elimina las imágenes temporales después de guardar el Excel
# for img_path in imagenes_generadas:
#     os.remove(img_path)
