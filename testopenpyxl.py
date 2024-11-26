# from openpyxl import Workbook
# wb = Workbook()

# # grab the active worksheet
# ws = wb.active

# # Data can be assigned directly to cells
# ws['A1'] = 42

# # Rows can also be appended
# ws.append([1, 2, 3, 4])

# # Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()

# # Save the file
# wb.save("sample.xlsx")


from openpyxl import load_workbook

# Ruta del archivo de Excel
excel_file = "sample.xlsx"

# Cargar el archivo de Excel
workbook = load_workbook(excel_file)

# Seleccionar la hoja activa (o una específica si se conoce su nombre)
sheet = workbook.active  # Esto selecciona la hoja que está activa por defecto
# o
# sheet = workbook["NombreDeLaHoja"]  # Esto selecciona una hoja específica por su nombre

# Leer datos de una celda específica
dato = sheet["A1"].value  # Lee el valor de la celda A1
print("Valor de A1:", dato)
